# Remplacez le début de votre fichier excel_generator.py par ceci :

import sys
import os
base_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
sys.path.append(base_path)

import pandas as pd
import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import json

# Import conditionnel du logger avec fallback
try:
    from utils.logger import CTILogger
except ImportError:
    print("⚠️ utils.logger non trouvé, utilisation du fallback")
    class CTILogger:
        def __init__(self, name):
            self.name = name
        
        def info(self, msg):
            print(f"[INFO] {self.name}: {msg}")
        
        def warning(self, msg):
            print(f"[WARNING] {self.name}: {msg}")
        
        def error(self, msg):
            print(f"[ERROR] {self.name}: {msg}")

# Import conditionnel de DatabaseManager (optionnel pour excel_generator)
try:
    from utils.database import DatabaseManager
    DATABASE_AVAILABLE = True
except ImportError:
    print("⚠️ utils.database non trouvé - Excel generator fonctionnera sans DB")
    DATABASE_AVAILABLE = False
    DatabaseManager = None

class ExcelGenerator:
    def __init__(self):
        self.logger = CTILogger("Excel_Generator")
        self.output_dir = "output/excel_reports"
        os.makedirs(self.output_dir, exist_ok=True)
    
    def _convert_to_excel_safe(self, value):
        """Convertit une valeur en format compatible Excel"""
        if value is None:
            return ''
        elif isinstance(value, list):
            return ', '.join(str(item) for item in value)
        elif isinstance(value, dict):
            return str(value)
        elif isinstance(value, bool):
            return 'Oui' if value else 'Non'
        else:
            return str(value)
    
    def _safe_load_json(self, file_path):
        """Charge un fichier JSON en gérant les problèmes d'encodage"""
        try:
            # Essayer d'abord avec utf-8
            with open(file_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except UnicodeDecodeError:
            try:
                # Si utf-8 échoue, essayer avec utf-8 et ignorer les erreurs
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    return json.load(f)
            except:
                try:
                    # Dernier recours : latin-1
                    with open(file_path, 'r', encoding='latin-1') as f:
                        return json.load(f)
                except:
                    self.logger.error(f"Impossible de lire le fichier {file_path}")
                    return None
        except json.JSONDecodeError as e:
            self.logger.error(f"Erreur JSON dans {file_path}: {e}")
            return None
        except Exception as e:
            self.logger.error(f"Erreur lors de la lecture de {file_path}: {e}")
            return None
    
    def _create_sample_data_if_missing(self):
        """Crée des données d'exemple si aucun fichier n'existe"""
        daily_feeds_dir = "output/daily_feeds"
        os.makedirs(daily_feeds_dir, exist_ok=True)
        
        today = datetime.now().strftime('%Y%m%d')
        
        # Données CVE d'exemple
        cve_file = f"{daily_feeds_dir}/critical_cves_{today}.json"
        if not os.path.exists(cve_file):
            sample_cves = [
                {
                    "cve_id": "CVE-2024-0001",
                    "published_date": "2024-01-15",
                    "cvss_score": 9.8,
                    "severity": "Critical",
                    "description": "Sample CVE for testing purposes",
                    "exploit_available": True,
                    "products": ["Product A", "Product B"]
                }
            ]
            try:
                with open(cve_file, 'w', encoding='utf-8') as f:
                    json.dump(sample_cves, f, ensure_ascii=False, indent=2)
                self.logger.info(f"Fichier d'exemple CVE créé : {cve_file}")
            except Exception as e:
                self.logger.error(f"Erreur création fichier CVE : {e}")
        
        # Données IOC d'exemple
        ioc_file = f"{daily_feeds_dir}/enriched_iocs_{today}.json"
        if not os.path.exists(ioc_file):
            sample_iocs = [
                {
                    "hash": "d41d8cd98f00b204e9800998ecf8427e",
                    "type": "md5",
                    "source": "Sample Source",
                    "enriched_at": "2024-01-15T10:00:00Z",
                    "detection_ratio": "5/65",
                    "malicious": 5,
                    "suspicious": 2,
                    "malware_families": ["Trojan", "Backdoor"],
                    "country": "Unknown"
                }
            ]
            try:
                with open(ioc_file, 'w', encoding='utf-8') as f:
                    json.dump(sample_iocs, f, ensure_ascii=False, indent=2)
                self.logger.info(f"Fichier d'exemple IOC créé : {ioc_file}")
            except Exception as e:
                self.logger.error(f"Erreur création fichier IOC : {e}")
        
        # Données OTX d'exemple
        otx_file = f"{daily_feeds_dir}/otx_pulses_{today}.json"
        if not os.path.exists(otx_file):
            sample_otx = [
                {
                    "name": "Sample Threat Intelligence",
                    "author_name": "Security Researcher",
                    "created": "2024-01-15T10:00:00Z",
                    "tags": ["malware", "apt"],
                    "industries": ["Technology", "Finance"],
                    "malware_families": ["Banking Trojan"],
                    "indicators": ["indicator1", "indicator2"]
                }
            ]
            try:
                with open(otx_file, 'w', encoding='utf-8') as f:
                    json.dump(sample_otx, f, ensure_ascii=False, indent=2)
                self.logger.info(f"Fichier d'exemple OTX créé : {otx_file}")
            except Exception as e:
                self.logger.error(f"Erreur création fichier OTX : {e}")
    
    def generate_cve_dashboard(self):
        """Génère le tableau de bord des CVE"""
        try:
            self.logger.info("Génération du tableau de bord CVE")
            
            # Créer des données d'exemple si nécessaire
            self._create_sample_data_if_missing()
            
            # Charger les données CVE
            cve_file = f"output/daily_feeds/critical_cves_{datetime.now().strftime('%Y%m%d')}.json"
            if not os.path.exists(cve_file):
                self.logger.warning("Aucune donnée CVE trouvée")
                return None
            
            cve_data = self._safe_load_json(cve_file)
            if not cve_data:
                self.logger.warning("Impossible de charger les données CVE")
                return None
            
            # Créer le DataFrame et convertir toutes les valeurs
            df = pd.DataFrame(cve_data)
            
            # Convertir toutes les valeurs en format compatible Excel
            for col in df.columns:
                df[col] = df[col].apply(self._convert_to_excel_safe)
            
            # Vérifier que les colonnes existent
            available_columns = df.columns.tolist()
            columns_mapping = {}
            
            # Mapping flexible des colonnes
            if 'cve_id' in available_columns:
                columns_mapping['cve_id'] = 'CVE ID'
            if 'published_date' in available_columns:
                columns_mapping['published_date'] = 'Date Publication'
            if 'cvss_score' in available_columns:
                columns_mapping['cvss_score'] = 'Score CVSS'
            if 'severity' in available_columns:
                columns_mapping['severity'] = 'Sévérité'
            if 'description' in available_columns:
                columns_mapping['description'] = 'Description'
            if 'exploit_available' in available_columns:
                columns_mapping['exploit_available'] = 'Exploit Public'
            if 'products' in available_columns:
                columns_mapping['products'] = 'Produits Impactés'
            
            if columns_mapping:
                df = df[list(columns_mapping.keys())].rename(columns=columns_mapping)
            
            # Créer le fichier Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "CVE Dashboard"
            
            # Ajouter les données
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Formater le header
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Ajuster la largeur des colonnes
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Sauvegarder
            filename = f"{self.output_dir}/CVE_Dashboard_{datetime.now().strftime('%Y%m%d')}.xlsx"
            wb.save(filename)
            
            self.logger.info(f"Tableau de bord CVE généré : {filename}")
            return filename
            
        except Exception as e:
            self.logger.error(f"Erreur lors de la génération du tableau CVE : {e}")
            return None
    
    def generate_ioc_tracking(self):
        """Génère le tableau de suivi des IOCs"""
        try:
            self.logger.info("Génération du tableau de suivi IOCs")
            
            # Charger les données IOC enrichies
            ioc_file = f"output/daily_feeds/enriched_iocs_{datetime.now().strftime('%Y%m%d')}.json"
            if not os.path.exists(ioc_file):
                self.logger.warning("Aucune donnée IOC enrichie trouvée")
                return None
            
            ioc_data = self._safe_load_json(ioc_file)
            if not ioc_data:
                self.logger.warning("Impossible de charger les données IOC")
                return None
            
            # Préparer les données pour le DataFrame avec conversion des listes
            formatted_data = []
            for ioc in ioc_data:
                # Convertir les listes en chaînes
                malware_families = ioc.get('malware_families', [])
                if isinstance(malware_families, list):
                    malware_families_str = ', '.join(malware_families)
                else:
                    malware_families_str = str(malware_families) if malware_families else ''
                
                formatted_data.append({
                    'IOC': ioc.get('hash', ioc.get('ip', ioc.get('domain', ioc.get('url', '')))),
                    'Type': ioc.get('type', ''),
                    'Source': ioc.get('source', ''),
                    'Date Détection': ioc.get('enriched_at', ''),
                    'Réputation VT': str(ioc.get('detection_ratio', 0)),
                    'Malicieux': ioc.get('malicious', 0),
                    'Suspect': ioc.get('suspicious', 0),
                    'Famille Malware': malware_families_str,
                    'Pays': ioc.get('country', ''),
                    'Statut': 'À analyser'
                })
            
            df = pd.DataFrame(formatted_data)
            
            # S'assurer que toutes les valeurs sont compatibles Excel
            for col in df.columns:
                df[col] = df[col].apply(self._convert_to_excel_safe)
            
            # S'assurer que toutes les valeurs sont compatibles Excel
            for col in df.columns:
                df[col] = df[col].apply(self._convert_to_excel_safe)
            
            # Créer le fichier Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "IOC Tracking"
            
            # Ajouter les données
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Formater
            self._format_excel_sheet(ws)
            
            # Sauvegarder
            filename = f"{self.output_dir}/IOC_Tracking_{datetime.now().strftime('%Y%m%d')}.xlsx"
            wb.save(filename)
            
            self.logger.info(f"Tableau de suivi IOC généré : {filename}")
            return filename
            
        except Exception as e:
            self.logger.error(f"Erreur lors de la génération du tableau IOC : {e}")
            return None
    
    def generate_threat_intelligence(self):
        """Génère le tableau de threat intelligence"""
        try:
            self.logger.info("Génération du tableau de threat intelligence")
            
            # Charger les données OTX
            otx_file = f"output/daily_feeds/otx_pulses_{datetime.now().strftime('%Y%m%d')}.json"
            if not os.path.exists(otx_file):
                self.logger.warning("Aucune donnée OTX trouvée")
                return None
            
            otx_data = self._safe_load_json(otx_file)
            if not otx_data:
                self.logger.warning("Impossible de charger les données OTX")
                return None
            
            # Préparer les données avec conversion des listes
            formatted_data = []
            for pulse in otx_data:
                # Convertir les listes en chaînes
                tags = pulse.get('tags', [])
                tags_str = ', '.join(tags) if isinstance(tags, list) else str(tags) if tags else ''
                
                industries = pulse.get('industries', [])
                industries_str = ', '.join(industries) if isinstance(industries, list) else str(industries) if industries else ''
                
                malware_families = pulse.get('malware_families', [])
                malware_families_str = ', '.join(malware_families) if isinstance(malware_families, list) else str(malware_families) if malware_families else ''
                
                indicators = pulse.get('indicators', [])
                indicators_count = len(indicators) if isinstance(indicators, list) else 0
                
                formatted_data.append({
                    'Menace': pulse.get('name', ''),
                    'Auteur': pulse.get('author_name', ''),
                    'Date': pulse.get('created', ''),
                    'Tags': tags_str,
                    'Industries': industries_str,
                    'Familles Malware': malware_families_str,
                    'Nb Indicateurs': indicators_count,
                    'Niveau Risque': self._calculate_risk_level(pulse),
                    'Statut': 'En cours d\'analyse'
                })
            
            df = pd.DataFrame(formatted_data)
            
            # Créer le fichier Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Threat Intelligence"
            
            # Ajouter les données
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Formater
            self._format_excel_sheet(ws)
            
            # Sauvegarder
            filename = f"{self.output_dir}/Threat_Intelligence_{datetime.now().strftime('%Y%m%d')}.xlsx"
            wb.save(filename)
            
            self.logger.info(f"Tableau de threat intelligence généré : {filename}")
            return filename
            
        except Exception as e:
            self.logger.error(f"Erreur lors de la génération du tableau TI : {e}")
            return None
    
    def generate_summary_report(self):
        """Génère un rapport de synthèse"""
        try:
            self.logger.info("Génération du rapport de synthèse")
            
            # Collecter les statistiques
            stats = self._collect_daily_stats()
            
            # Créer le workbook avec plusieurs feuilles
            wb = Workbook()
            
            # Feuille de synthèse
            ws_summary = wb.active
            ws_summary.title = "Synthèse Quotidienne"
            
            # Ajouter les statistiques
            summary_data = [
                ['Métriques', 'Valeur'],
                ['CVE Critiques', stats.get('cve_count', 0)],
                ['IOCs Enrichis', stats.get('ioc_count', 0)],
                ['Pulses OTX', stats.get('pulse_count', 0)],
                ['Feeds Collectés', stats.get('feed_count', 0)],
                ['Dernière Mise à Jour', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
            ]
            
            for row in summary_data:
                ws_summary.append(row)
            
            # Formater
            self._format_excel_sheet(ws_summary)
            
            # Sauvegarder
            filename = f"{self.output_dir}/Summary_Report_{datetime.now().strftime('%Y%m%d')}.xlsx"
            wb.save(filename)
            
            self.logger.info(f"Rapport de synthèse généré : {filename}")
            return filename
            
        except Exception as e:
            self.logger.error(f"Erreur lors de la génération du rapport de synthèse : {e}")
            return None
    
    def _format_excel_sheet(self, worksheet):
        """Formate une feuille Excel"""
        try:
            # Formater le header
            if worksheet.max_row > 0:
                for cell in worksheet[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
            
            # Ajuster la largeur des colonnes
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        except Exception as e:
            self.logger.error(f"Erreur lors du formatage Excel : {e}")
    
    def _calculate_risk_level(self, pulse):
        """Calcule le niveau de risque d'un pulse"""
        try:
            risk_score = 0
            
            # Basé sur les tags
            high_risk_tags = ['apt', 'ransomware', '0day', 'exploit', 'malware']
            medium_risk_tags = ['phishing', 'trojan', 'botnet']
            
            tags = pulse.get('tags', [])
            if isinstance(tags, list):
                tags = [tag.lower() for tag in tags]
            else:
                tags = [str(tags).lower()] if tags else []
            
            for tag in tags:
                if any(high_tag in tag for high_tag in high_risk_tags):
                    risk_score += 3
                elif any(med_tag in tag for med_tag in medium_risk_tags):
                    risk_score += 2
                else:
                    risk_score += 1
            
            # Basé sur le nombre d'indicateurs
            indicators = pulse.get('indicators', [])
            ioc_count = len(indicators) if isinstance(indicators, list) else 0
            
            if ioc_count > 10:
                risk_score += 2
            elif ioc_count > 5:
                risk_score += 1
            
            # Déterminer le niveau
            if risk_score >= 8:
                return 'Critique'
            elif risk_score >= 5:
                return 'Élevé'
            elif risk_score >= 3:
                return 'Moyen'
            else:
                return 'Faible'
        except Exception:
            return 'Non défini'
    
    def _collect_daily_stats(self):
        """Collecte les statistiques quotidiennes"""
        stats = {}
        
        try:
            # Statistiques CVE
            cve_file = f"output/daily_feeds/critical_cves_{datetime.now().strftime('%Y%m%d')}.json"
            if os.path.exists(cve_file):
                cve_data = self._safe_load_json(cve_file)
                stats['cve_count'] = len(cve_data) if cve_data else 0
            else:
                stats['cve_count'] = 0
            
            # Statistiques IOC
            ioc_file = f"output/daily_feeds/enriched_iocs_{datetime.now().strftime('%Y%m%d')}.json"
            if os.path.exists(ioc_file):
                ioc_data = self._safe_load_json(ioc_file)
                stats['ioc_count'] = len(ioc_data) if ioc_data else 0
            else:
                stats['ioc_count'] = 0
            
            # Statistiques OTX
            otx_file = f"output/daily_feeds/otx_pulses_{datetime.now().strftime('%Y%m%d')}.json"
            if os.path.exists(otx_file):
                otx_data = self._safe_load_json(otx_file)
                stats['pulse_count'] = len(otx_data) if otx_data else 0
            else:
                stats['pulse_count'] = 0
            
            # Statistiques feeds
            feed_file = f"output/daily_feeds/daily_summary_{datetime.now().strftime('%Y%m%d')}.json"
            if os.path.exists(feed_file):
                feed_data = self._safe_load_json(feed_file)
                stats['feed_count'] = feed_data.get('total_items', 0) if feed_data else 0
            else:
                stats['feed_count'] = 0
            
        except Exception as e:
            self.logger.error(f"Erreur lors de la collecte des statistiques : {e}")
            # Définir des valeurs par défaut en cas d'erreur
            stats = {
                'cve_count': 0,
                'ioc_count': 0,
                'pulse_count': 0,
                'feed_count': 0
            }
        
        return stats

# Test du générateur
if __name__ == "__main__":
    generator = ExcelGenerator()
    
    # Générer tous les rapports
    generator.generate_cve_dashboard()
    generator.generate_ioc_tracking()
    generator.generate_threat_intelligence()
    generator.generate_summary_report()